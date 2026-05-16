import os
import uuid
import shutil
import asyncio
import logging
from pathlib import Path
from contextlib import asynccontextmanager
from typing import Optional

from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# ─── Logging ────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ─── Config ─────────────────────────────────────────────────────────────────
TEMP_DIR = Path("/tmp/pdf_converter")
TEMP_DIR.mkdir(parents=True, exist_ok=True)
MAX_FILE_SIZE_MB = 50
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
ALLOWED_MIME = {"application/pdf", "application/octet-stream"}
CLEANUP_AFTER_SECONDS = 300  # 5 minutes


# ─── Lifespan: background cleanup task ────────────────────────────────────── sigma boys only here
@asynccontextmanager
async def lifespan(app: FastAPI):
    task = asyncio.create_task(cleanup_loop())
    yield
    task.cancel()


async def cleanup_loop():
    """Delete temp files older than CLEANUP_AFTER_SECONDS every minute."""
    while True:
        await asyncio.sleep(60)
        try:
            for f in TEMP_DIR.iterdir():
                age = asyncio.get_event_loop().time() - f.stat().st_mtime
                if age > CLEANUP_AFTER_SECONDS:
                    f.unlink(missing_ok=True)
                    logger.info(f"Cleaned up {f.name}")
        except Exception as e:
            logger.error(f"Cleanup error: {e}")


# ─── Rate limiter ────────────────────────────────────────────────────────────
limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="PDF Converter API",
    description="Convert PDF files to Word (.docx) or Excel (.xlsx) — 100% free, no storage.",
    version="1.0.0",
    lifespan=lifespan,
)

app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # tighten in prod: ["https://yourdomain.com"]
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Helpers ─────────────────────────────────────────────────────────────────
def validate_pdf(file: UploadFile, content: bytes) -> None:
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only .pdf files are accepted.")
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise HTTPException(413, f"File too large. Maximum size is {MAX_FILE_SIZE_MB} MB.")
    if not content.startswith(b"%PDF"):
        raise HTTPException(400, "Invalid PDF file (bad magic bytes).")


def temp_path(ext: str) -> Path:
    return TEMP_DIR / f"{uuid.uuid4()}.{ext}"


# ─── Routes ──────────────────────────────────────────────────────────────────
@app.get("/health")
async def health():
    return {"status": "ok"}

@app.get("/debug-env")
async def debug_env():
    key = os.environ.get("GEMINI_API_KEY", "")
    return {
        "key_present": bool(key),
        "key_length": len(key),
        "key_prefix": key[:8] if key else None
    }


@app.post("/convert/word")
@limiter.limit("10/minute")
async def convert_to_word(request: Request, file: UploadFile = File(...)):
    """Convert PDF → DOCX using pdf2docx (free, offline)."""
    from pdf2docx import Converter

    content = await file.read()
    validate_pdf(file, content)

    pdf_path = temp_path("pdf")
    docx_path = temp_path("docx")

    try:
        pdf_path.write_bytes(content)
        logger.info(f"Converting {file.filename} → DOCX")

        cv = Converter(str(pdf_path))
        cv.convert(str(docx_path), start=0, end=None)
        cv.close()

        if not docx_path.exists() or docx_path.stat().st_size == 0:
            raise HTTPException(500, "Conversion produced an empty file.")

        stem = Path(file.filename).stem
        return FileResponse(
            path=str(docx_path),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename=f"{stem}.docx",
            background=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"DOCX conversion failed: {e}")
        raise HTTPException(500, f"Conversion failed: {str(e)}")
    finally:
        pdf_path.unlink(missing_ok=True)
        # docx cleaned up by background task after 5 min


@app.post("/convert/excel")
@limiter.limit("10/minute")
async def convert_to_excel(request: Request, file: UploadFile = File(...)):
    """Convert PDF → XLSX by extracting tables with pdfplumber (free, offline)."""
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    content = await file.read()
    validate_pdf(file, content)

    pdf_path = temp_path("pdf")
    xlsx_path = temp_path("xlsx")

    try:
        pdf_path.write_bytes(content)
        logger.info(f"Converting {file.filename} → XLSX")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1a1a2e")
        alt_fill = PatternFill("solid", fgColor="F2F2F7")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        table_count = 0

        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()

                # Also try to extract text as a fallback sheet for page 1
                if page_num == 1 and not tables:
                    text = page.extract_text()
                    if text:
                        ws = wb.create_sheet(title="Text Content")
                        for i, line in enumerate(text.split("\n"), start=1):
                            ws.cell(row=i, column=1, value=line)
                        ws.column_dimensions["A"].width = 80
                        continue

                for t_idx, table in enumerate(tables):
                    if not table or all(all(c is None for c in row) for row in table):
                        continue

                    table_count += 1
                    sheet_name = f"P{page_num}_T{t_idx+1}" if len(tables) > 1 else f"Page {page_num}"
                    # Excel sheet names max 31 chars
                    ws = wb.create_sheet(title=sheet_name[:31])

                    for r_idx, row in enumerate(table, start=1):
                        for c_idx, cell_val in enumerate(row, start=1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=cell_val or "")
                            cell.border = border
                            cell.alignment = center
                            if r_idx == 1:
                                cell.font = header_font
                                cell.fill = header_fill
                            elif r_idx % 2 == 0:
                                cell.fill = alt_fill

                    # Auto-fit columns (approximate)
                    for col in ws.columns:
                        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

                    ws.row_dimensions[1].height = 22

        if not wb.sheetnames:
            # No tables found — create info sheet
            ws = wb.create_sheet(title="No Tables Found")
            ws["A1"] = "No tables were detected in this PDF."
            ws["A2"] = "Try the Word conversion to preserve text layout."
            ws.column_dimensions["A"].width = 60

        wb.save(str(xlsx_path))

        stem = Path(file.filename).stem
        return FileResponse(
            path=str(xlsx_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{stem}.xlsx",
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"XLSX conversion failed: {e}")
        raise HTTPException(500, f"Conversion failed: {str(e)}")
    finally:
        pdf_path.unlink(missing_ok=True)


# ─── UPI Tracker Endpoint ────────────────────────────────────────────────────
@app.post("/convert/upi-tracker")
@limiter.limit("5/minute")
async def convert_upi_tracker(request: Request, file: UploadFile = File(...)):
    """
    Extract outgoing UPI payments from a bank statement PDF,
    categorize them (Food, Transport, etc.), and return a
    formatted Excel with category → merchant → amount breakdown.
    """
    from google import genai
    from google.genai import types
    import json
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
    if not GEMINI_API_KEY:
        raise HTTPException(500, "GEMINI_API_KEY not configured on server.")

    content = await file.read()
    validate_pdf(file, content)

    pdf_path  = temp_path("pdf")
    xlsx_path = temp_path("xlsx")

    try:
        pdf_path.write_bytes(content)
        logger.info(f"UPI tracker: processing {file.filename}")

        # ── Step 1: Extract text from PDF ────────────────────────────────────
        raw_text = ""
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    raw_text += page_text + "\n"

        # ── Step 2: Send to Gemini ────────────────────────────────────────────
        client = genai.Client(api_key=GEMINI_API_KEY)
        logger.info("Sending to Gemini for UPI extraction...")

        if len(raw_text.strip()) > 200:
            # Text-based PDF — send extracted text (faster, cheaper)
            response = client.models.generate_content(
                model="gemini-1.5-flash-latest",
                contents=UPI_PROMPT + "\n\nBANK STATEMENT TEXT:\n" + raw_text
            )
        else:
            # Scanned/image PDF — send PDF bytes directly for vision
            response = client.models.generate_content(
                model="gemini-1.5-flash-latest",
                contents=[
                    types.Part.from_bytes(data=content, mime_type="application/pdf"),
                    UPI_PROMPT
                ]
            )

        raw = response.text.strip()

        # Strip markdown fences if present
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        raw = raw.strip()

        data = json.loads(raw)

        transactions     = data.get("transactions", [])
        bank_name        = data.get("bank_name", "")
        account_holder   = data.get("account_holder", "")
        statement_period = data.get("statement_period", "")

        if not transactions:
            raise HTTPException(422, "No outgoing UPI payments found in this statement.")

        # ── Step 3: Group by category → merchant ─────────────────────────────
        grouped = defaultdict(lambda: defaultdict(list))
        for txn in transactions:
            cat      = str(txn.get("category", "Other")).strip() or "Other"
            merchant = str(txn.get("merchant", "Unknown")).strip() or "Unknown"
            try:
                amount = float(str(txn.get("amount", 0)).replace(",", ""))
            except (ValueError, TypeError):
                amount = 0.0
            grouped[cat][merchant].append(amount)

        # ── Step 4: Build Excel ───────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UPI Spending"

        # Styles
        meta_font     = Font(bold=True, size=11, color="1e40af")
        cat_font      = Font(bold=True, size=12, color="FFFFFF")
        cat_fills     = [
            PatternFill("solid", fgColor="1e40af"),
            PatternFill("solid", fgColor="0f766e"),
            PatternFill("solid", fgColor="7c3aed"),
            PatternFill("solid", fgColor="b45309"),
            PatternFill("solid", fgColor="be123c"),
            PatternFill("solid", fgColor="0369a1"),
            PatternFill("solid", fgColor="4d7c0f"),
            PatternFill("solid", fgColor="9333ea"),
        ]
        merchant_font = Font(size=11, color="1e293b")
        merchant_fill = PatternFill("solid", fgColor="F1F5F9")
        alt_fill      = PatternFill("solid", fgColor="FFFFFF")
        sub_font      = Font(bold=True, size=11, color="475569")
        sub_fill      = PatternFill("solid", fgColor="E2E8F0")
        grand_font    = Font(bold=True, size=12, color="FFFFFF")
        grand_fill    = PatternFill("solid", fgColor="0f172a")
        thin          = Side(style="thin", color="CBD5E1")
        border        = Border(left=thin, right=thin, top=thin, bottom=thin)
        left_align    = Alignment(horizontal="left",  vertical="center", indent=1)
        right_align   = Alignment(horizontal="right", vertical="center")
        center_align  = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 18

        # Meta rows
        current_row = 1
        for label, value in [("Bank", bank_name), ("Account Holder", account_holder), ("Statement Period", statement_period)]:
            ws.cell(row=current_row, column=1, value=label).font = meta_font
            ws.cell(row=current_row, column=2, value=value).alignment = right_align
            current_row += 1

        current_row += 1  # blank row

        # Column headers
        hdr_fill = PatternFill("solid", fgColor="334155")
        for col, label in [(1, "Service / Merchant"), (2, "Amount (₹)")]:
            cell = ws.cell(row=current_row, column=col, value=label)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = hdr_fill
            cell.border = border
            cell.alignment = center_align
        ws.row_dimensions[current_row].height = 22
        ws.freeze_panes = ws.cell(row=current_row + 1, column=1)
        current_row += 1

        # Category blocks
        grand_total = 0.0
        CATEGORY_EMOJIS = {
            "food": "🍔", "transport": "🚗", "travel": "✈️",
            "shopping": "🛍️", "entertainment": "🎬", "subscriptions": "📱",
            "medicine": "💊", "health": "💊", "education": "📚",
            "utilities": "💡", "bills": "📄", "groceries": "🛒",
            "salary": "💰", "transfer": "🔄", "other": "📦",
        }

        def get_emoji(cat_name: str) -> str:
            return CATEGORY_EMOJIS.get(cat_name.lower(), "📦")

        for cat_idx, (category, merchants) in enumerate(sorted(grouped.items())):
            fill  = cat_fills[cat_idx % len(cat_fills)]
            emoji = get_emoji(category)

            cell_a = ws.cell(row=current_row, column=1, value=f"  {emoji}  {category.upper()}")
            cell_a.font = cat_font; cell_a.fill = fill; cell_a.border = border; cell_a.alignment = left_align
            cell_b = ws.cell(row=current_row, column=2, value="")
            cell_b.fill = fill; cell_b.border = border
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            cat_total    = 0.0
            merchant_alt = False

            for merchant, amounts in sorted(merchants.items()):
                merchant_total = round(sum(amounts), 2)
                cat_total     += merchant_total
                mfill          = merchant_fill if merchant_alt else alt_fill
                merchant_alt   = not merchant_alt

                cell_a = ws.cell(row=current_row, column=1, value=f"    {merchant}")
                cell_a.font = merchant_font; cell_a.fill = mfill; cell_a.border = border; cell_a.alignment = left_align

                cell_b = ws.cell(row=current_row, column=2, value=merchant_total)
                cell_b.font = merchant_font; cell_b.fill = mfill; cell_b.border = border
                cell_b.alignment = right_align; cell_b.number_format = '#,##0.00'
                current_row += 1

            grand_total += cat_total
            cell_a = ws.cell(row=current_row, column=1, value=f"  Subtotal — {category}")
            cell_a.font = sub_font; cell_a.fill = sub_fill; cell_a.border = border; cell_a.alignment = left_align
            cell_b = ws.cell(row=current_row, column=2, value=round(cat_total, 2))
            cell_b.font = sub_font; cell_b.fill = sub_fill; cell_b.border = border
            cell_b.alignment = right_align; cell_b.number_format = '#,##0.00'
            current_row += 2

        # Grand Total
        cell_a = ws.cell(row=current_row, column=1, value="  💳  GRAND TOTAL")
        cell_a.font = grand_font; cell_a.fill = grand_fill; cell_a.border = border; cell_a.alignment = left_align
        cell_b = ws.cell(row=current_row, column=2, value=round(grand_total, 2))
        cell_b.font = grand_font; cell_b.fill = grand_fill; cell_b.border = border
        cell_b.alignment = right_align; cell_b.number_format = '#,##0.00'
        ws.row_dimensions[current_row].height = 24

        # Raw Transactions sheet
        ws2 = wb.create_sheet(title="Raw Transactions")
        hdr2_fill = PatternFill("solid", fgColor="1e40af")
        for col, h in enumerate(["Date", "Merchant", "Category", "Amount (₹)", "UPI ID / Reference"], start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = hdr2_fill; cell.border = border; cell.alignment = center_align
        ws2.row_dimensions[1].height = 20

        for i, txn in enumerate(transactions, start=2):
            try:
                amt = float(str(txn.get("amount", 0)).replace(",", ""))
            except (ValueError, TypeError):
                amt = 0.0
            row_fill = merchant_fill if i % 2 == 0 else alt_fill
            for col, val in enumerate([txn.get("date",""), txn.get("merchant",""), txn.get("category",""), amt, txn.get("upi_id","")], start=1):
                cell = ws2.cell(row=i, column=col, value=val)
                cell.fill = row_fill; cell.border = border; cell.alignment = left_align
                if col == 4:
                    cell.number_format = '#,##0.00'

        for col_letter, width in zip(["A","B","C","D","E"], [14, 28, 18, 16, 32]):
            ws2.column_dimensions[col_letter].width = width
        ws2.freeze_panes = ws2["A2"]

        wb.save(str(xlsx_path))

        stem = Path(file.filename).stem
        return FileResponse(
            path=str(xlsx_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{stem}_upi_spending.xlsx",
        )

    except HTTPException:
        raise
    except json.JSONDecodeError:
        logger.error("Gemini returned non-JSON for UPI tracker")
        raise HTTPException(500, "Could not parse AI response. Try a clearer PDF.")
    except Exception as e:
        logger.error(f"UPI tracker error: {e}")
        raise HTTPException(500, f"Processing failed: {str(e)}")
    finally:
        pdf_path.unlink(missing_ok=True)
        # xlsx_path cleaned by background task


UPI_PROMPT = """You are a bank statement analyst. Your job is to extract ONLY outgoing UPI payments from the statement.

RULES:
- Include ONLY debit/outgoing UPI transactions (payments made BY the account holder)
- EXCLUDE: incoming credits, salary, refunds, ATM withdrawals, NEFT/RTGS/IMPS transfers to individuals, bank charges
- UPI transactions are typically described with UPI/, UPI-pay, UPI ref, or contain a UPI ID like merchant@bank
- Identify the real merchant name from the UPI ID or description (e.g. "ZOMATO@icici" → "Zomato", "SWIGGY.IN@axisb" → "Swiggy")
- Assign each transaction a category from this list ONLY: Food, Transport, Shopping, Entertainment, Subscriptions, Medicine, Groceries, Utilities, Education, Other
- Common mappings: Zomato/Swiggy/Blinkit/EatSure → Food, Uber/Ola/Rapido → Transport, Netflix/Spotify/Amazon Prime → Subscriptions, Amazon/Flipkart/Myntra → Shopping, Pharmeasy/Apollo/Netmeds → Medicine, BigBasket/Zepto/JioMart → Groceries, Electricity/Gas/Water bill → Utilities

Return ONLY a valid JSON object, no markdown, no explanation:

{
  "bank_name": "name of the bank",
  "account_holder": "account holder name if visible",
  "statement_period": "e.g. April 2025",
  "transactions": [
    {
      "date": "DD/MM/YYYY",
      "merchant": "Clean merchant name e.g. Zomato",
      "category": "Food",
      "amount": 450.00,
      "upi_id": "raw UPI ID or reference from statement"
    }
  ]
}

If no outgoing UPI transactions are found, return an empty transactions array.
Return ONLY the JSON, nothing else."""
