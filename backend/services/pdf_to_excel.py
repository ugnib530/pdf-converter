"""
services/pdf_to_excel.py
Extracts tables from a PDF with pdfplumber and writes them to an
Excel workbook (.xlsx) using openpyxl.

Called by routers/from_pdf.py → pdf_to_excel endpoint.
"""
import asyncio
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def _do_convert(pdf_path: Path, xlsx_path: Path) -> None:
    """Synchronous extraction — runs inside run_in_executor."""
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── Style constants ───────────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1a1a2e")
    alt_fill     = PatternFill("solid", fgColor="F2F2F7")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()

            # Fallback: if no tables on page 1, write the raw text
            if page_num == 1 and not tables:
                text = page.extract_text()
                if text:
                    ws = wb.create_sheet(title="Text Content")
                    for i, line in enumerate(text.split("\n"), start=1):
                        ws.cell(row=i, column=1, value=line)
                    ws.column_dimensions["A"].width = 80
                continue

            for t_idx, table in enumerate(tables):
                if not table or all(all(c is None for c in row) for row in table):
                    continue

                sheet_name = (
                    f"P{page_num}_T{t_idx + 1}" if len(tables) > 1 else f"Page {page_num}"
                )
                ws = wb.create_sheet(title=sheet_name[:31])

                for r_idx, row in enumerate(table, start=1):
                    for c_idx, cell_val in enumerate(row, start=1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=cell_val or "")
                        cell.border = border
                        cell.alignment = center_align
                        if r_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                        elif r_idx % 2 == 0:
                            cell.fill = alt_fill

                # Auto-fit columns (approximate — openpyxl has no auto-fit)
                for col in ws.columns:
                    max_len = max(
                        (len(str(c.value)) for c in col if c.value), default=10
                    )
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
                ws.row_dimensions[1].height = 22

    if not wb.sheetnames:
        ws = wb.create_sheet(title="No Tables Found")
        ws["A1"] = "No tables were detected in this PDF."
        ws["A2"] = "Try the Word conversion to preserve text layout."
        ws.column_dimensions["A"].width = 60

    wb.save(str(xlsx_path))


async def convert_pdf_to_excel(pdf_path: Path, xlsx_path: Path) -> None:
    """
    Async wrapper for the synchronous pdfplumber/openpyxl work.

    Args:
        pdf_path:  Source PDF (must exist on disk).
        xlsx_path: Destination XLSX path.
    """
    loop = asyncio.get_event_loop()
    logger.info(f"Starting PDF→XLSX: {pdf_path.name}")
    await loop.run_in_executor(None, _do_convert, pdf_path, xlsx_path)
    logger.info(f"PDF→XLSX complete: {xlsx_path.stat().st_size:,} bytes")
