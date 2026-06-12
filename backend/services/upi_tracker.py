"""
services/upi_tracker.py
Extracts outgoing UPI payments from a bank statement PDF, categorises them
via Gemini, and writes a formatted Excel workbook.

Called by routers/from_pdf.py → upi_tracker endpoint.
"""
import asyncio
import json
import logging
from collections import defaultdict
from pathlib import Path

logger = logging.getLogger(__name__)

# ── Prompt (unchanged from original main.py) ──────────────────────────────────
_UPI_PROMPT = """\
You are a bank statement analyst. Your job is to extract ONLY outgoing UPI payments.

RULES:
- Include ONLY debit/outgoing UPI transactions (payments made BY the account holder)
- EXCLUDE: incoming credits, salary, refunds, ATM withdrawals, NEFT/RTGS/IMPS transfers
  to individuals, bank charges
- UPI transactions are typically described with UPI/, UPI-pay, UPI ref, or contain a
  UPI ID like merchant@bank
- Identify the real merchant name from the UPI ID or description
  (e.g. "ZOMATO@icici" → "Zomato", "SWIGGY.IN@axisb" → "Swiggy")
- Assign each transaction a category from this list ONLY:
  Food, Transport, Shopping, Entertainment, Subscriptions, Medicine,
  Groceries, Utilities, Education, Other
- Common mappings: Zomato/Swiggy/Blinkit → Food, Uber/Ola/Rapido → Transport,
  Netflix/Spotify/Amazon Prime → Subscriptions, Amazon/Flipkart/Myntra → Shopping,
  Pharmeasy/Apollo/Netmeds → Medicine, BigBasket/Zepto/JioMart → Groceries,
  Electricity/Gas/Water → Utilities

Return ONLY a valid JSON object, no markdown, no explanation:

{
  "bank_name": "name of the bank",
  "account_holder": "account holder name if visible",
  "statement_period": "e.g. April 2025",
  "transactions": [
    {
      "date": "DD/MM/YYYY",
      "merchant": "Clean merchant name e.g. Zomato",
      "category": "Food",
      "amount": 450.00,
      "upi_id": "raw UPI ID or reference from statement"
    }
  ]
}

If no outgoing UPI transactions are found, return an empty transactions array.
Return ONLY the JSON, nothing else.\
"""

# ── Category emoji map ────────────────────────────────────────────────────────
_CATEGORY_EMOJIS = {
    "food": "🍔", "transport": "🚗", "travel": "✈️",
    "shopping": "🛍️", "entertainment": "🎬", "subscriptions": "📱",
    "medicine": "💊", "health": "💊", "education": "📚",
    "utilities": "💡", "bills": "📄", "groceries": "🛒",
    "salary": "💰", "transfer": "🔄", "other": "📦",
}


def _emoji(cat: str) -> str:
    return _CATEGORY_EMOJIS.get(cat.lower(), "📦")


# ── Synchronous core (runs in executor) ───────────────────────────────────────
def _do_run(pdf_path: Path, xlsx_path: Path, pdf_bytes: bytes, api_key: str) -> None:
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from google import genai
    from google.genai import types

    # ── 1. Extract text ───────────────────────────────────────────────────────
    raw_text = ""
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                raw_text += t + "\n"

    # ── 2. Call Gemini ────────────────────────────────────────────────────────
    client = genai.Client(api_key=api_key)
    if len(raw_text.strip()) > 200:
        response = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=_UPI_PROMPT + "\n\nBANK STATEMENT TEXT:\n" + raw_text,
        )
    else:
        response = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=[
                types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"),
                _UPI_PROMPT,
            ],
        )

    raw = response.text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()

    data = json.loads(raw)
    transactions   = data.get("transactions", [])
    bank_name      = data.get("bank_name", "")
    account_holder = data.get("account_holder", "")
    period         = data.get("statement_period", "")

    if not transactions:
        from fastapi import HTTPException
        raise HTTPException(422, "No outgoing UPI payments found in this statement.")

    # ── 3. Group by category → merchant ──────────────────────────────────────
    grouped: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for txn in transactions:
        cat      = str(txn.get("category", "Other")).strip() or "Other"
        merchant = str(txn.get("merchant", "Unknown")).strip() or "Unknown"
        try:
            amount = float(str(txn.get("amount", 0)).replace(",", ""))
        except (ValueError, TypeError):
            amount = 0.0
        grouped[cat][merchant].append(amount)

    # ── 4. Build Excel ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UPI Spending"

    # Styles
    meta_font    = Font(bold=True, size=11, color="1e40af")
    cat_font     = Font(bold=True, size=12, color="FFFFFF")
    cat_fills    = [
        PatternFill("solid", fgColor=c) for c in [
            "1e40af","0f766e","7c3aed","b45309",
            "be123c","0369a1","4d7c0f","9333ea",
        ]
    ]
    m_font  = Font(size=11, color="1e293b")
    m_fill  = PatternFill("solid", fgColor="F1F5F9")
    alt_fill = PatternFill("solid", fgColor="FFFFFF")
    sub_font = Font(bold=True, size=11, color="475569")
    sub_fill = PatternFill("solid", fgColor="E2E8F0")
    g_font   = Font(bold=True, size=12, color="FFFFFF")
    g_fill   = PatternFill("solid", fgColor="0f172a")
    thin     = Side(style="thin", color="CBD5E1")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    l_align  = Alignment(horizontal="left",   vertical="center", indent=1)
    r_align  = Alignment(horizontal="right",  vertical="center")
    c_align  = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18

    row = 1
    for label, value in [
        ("Bank", bank_name),
        ("Account Holder", account_holder),
        ("Statement Period", period),
    ]:
        ws.cell(row=row, column=1, value=label).font = meta_font
        ws.cell(row=row, column=2, value=value).alignment = r_align
        row += 1

    row += 1  # blank

    hdr_fill = PatternFill("solid", fgColor="334155")
    for col, lbl in [(1, "Service / Merchant"), (2, "Amount (₹)")]:
        c = ws.cell(row=row, column=col, value=lbl)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = hdr_fill; c.border = border; c.alignment = c_align
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    row += 1

    grand_total = 0.0
    for cat_idx, (category, merchants) in enumerate(sorted(grouped.items())):
        fill = cat_fills[cat_idx % len(cat_fills)]
        emoji = _emoji(category)

        ca = ws.cell(row=row, column=1, value=f"  {emoji}  {category.upper()}")
        ca.font = cat_font; ca.fill = fill; ca.border = border; ca.alignment = l_align
        cb = ws.cell(row=row, column=2, value="")
        cb.fill = fill; cb.border = border
        ws.row_dimensions[row].height = 22
        row += 1

        cat_total = 0.0
        alt = False
        for merchant, amounts in sorted(merchants.items()):
            total = round(sum(amounts), 2)
            cat_total += total
            mf = m_fill if alt else alt_fill
            alt = not alt

            ma = ws.cell(row=row, column=1, value=f"    {merchant}")
            ma.font = m_font; ma.fill = mf; ma.border = border; ma.alignment = l_align
            mb = ws.cell(row=row, column=2, value=total)
            mb.font = m_font; mb.fill = mf; mb.border = border
            mb.alignment = r_align; mb.number_format = "#,##0.00"
            row += 1

        grand_total += cat_total
        sa = ws.cell(row=row, column=1, value=f"  Subtotal — {category}")
        sa.font = sub_font; sa.fill = sub_fill; sa.border = border; sa.alignment = l_align
        sb = ws.cell(row=row, column=2, value=round(cat_total, 2))
        sb.font = sub_font; sb.fill = sub_fill; sb.border = border
        sb.alignment = r_align; sb.number_format = "#,##0.00"
        row += 2

    ga = ws.cell(row=row, column=1, value="  💳  GRAND TOTAL")
    ga.font = g_font; ga.fill = g_fill; ga.border = border; ga.alignment = l_align
    gb = ws.cell(row=row, column=2, value=round(grand_total, 2))
    gb.font = g_font; gb.fill = g_fill; gb.border = border
    gb.alignment = r_align; gb.number_format = "#,##0.00"
    ws.row_dimensions[row].height = 24

    # Raw Transactions sheet
    ws2 = wb.create_sheet(title="Raw Transactions")
    hdr2 = PatternFill("solid", fgColor="1e40af")
    for col, h in enumerate(
        ["Date", "Merchant", "Category", "Amount (₹)", "UPI ID / Reference"], start=1
    ):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = hdr2; c.border = border; c.alignment = c_align
    ws2.row_dimensions[1].height = 20

    for i, txn in enumerate(transactions, start=2):
        try:
            amt = float(str(txn.get("amount", 0)).replace(",", ""))
        except (ValueError, TypeError):
            amt = 0.0
        rf = m_fill if i % 2 == 0 else alt_fill
        for col, val in enumerate(
            [txn.get("date",""), txn.get("merchant",""), txn.get("category",""),
             amt, txn.get("upi_id","")],
            start=1,
        ):
            c = ws2.cell(row=i, column=col, value=val)
            c.fill = rf; c.border = border; c.alignment = l_align
            if col == 4:
                c.number_format = "#,##0.00"

    for col_letter, width in zip(["A","B","C","D","E"], [14, 28, 18, 16, 32]):
        ws2.column_dimensions[col_letter].width = width
    ws2.freeze_panes = ws2["A2"]

    wb.save(str(xlsx_path))


# ── Public async interface ────────────────────────────────────────────────────
async def run_upi_tracker(
    pdf_path: Path, xlsx_path: Path, pdf_bytes: bytes, api_key: str
) -> None:
    """
    Async wrapper — offloads synchronous work to a thread pool.

    Args:
        pdf_path:  Source PDF (must exist on disk).
        xlsx_path: Destination XLSX.
        pdf_bytes: Raw PDF bytes (sent to Gemini vision if text extraction fails).
        api_key:   Gemini API key.
    """
    loop = asyncio.get_event_loop()
    logger.info(f"Starting UPI Tracker: {pdf_path.name}")
    await loop.run_in_executor(None, _do_run, pdf_path, xlsx_path, pdf_bytes, api_key)
    logger.info(f"UPI Tracker complete: {xlsx_path.stat().st_size:,} bytes")
